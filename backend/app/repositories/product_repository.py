from __future__ import annotations

import hashlib
import json
import re
from pathlib import Path
from zipfile import ZipFile
from xml.etree import ElementTree as ET
from typing import Any

from openpyxl import load_workbook

from app.core.config import Settings


FIELD_ALIASES = {
    "系统款式": "系统款式",
    "系统\n款式": "系统款式",
    "商品名称": "商品名称",
    "商品编码": "商品编码",
    "商品图片": "商品图片",
    "商品二维码": "商品二维码",
    "产品优势卖点": "产品优势卖点",
    "主材质": "主材质",
    "配石材质": "配石材质",
    "开团价": "开团价",
    "批发裸价": "批发裸价",
    "折扣": "折扣",
    "系统属性": "系统属性",
    "适合人群": "适合人群",
    "显贵款": "显贵款",
}


def _normalize_header(value: Any) -> str:
    raw = str(value or "").replace("\n", "").replace(" ", "").strip()
    return FIELD_ALIASES.get(raw, raw)


def _safe_float(value: Any) -> float | None:
    if value in (None, ""):
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def _clean_text(value: Any) -> str:
    return re.sub(r"\s+", " ", str(value or "")).strip()


def _extract_dispimg_id(value: Any) -> str | None:
    text = str(value or "").strip()
    if not text:
        return None
    match = re.search(r'DISPIMG\("([^"]+)"', text, re.IGNORECASE)
    if match:
        return match.group(1)
    return None


class ProductRepository:
    def __init__(self, settings: Settings) -> None:
        self.settings = settings
        self.workbook_path: Path | None = None
        self.catalog_source_path: Path | None = None
        self.products: list[dict[str, Any]] = []
        self.loaded = False
        self.asset_url_by_image_id: dict[str, str] = {}

    def _discover_json_path(self) -> Path | None:
        if self.settings.product_json_path:
            path = Path(self.settings.product_json_path)
            if path.exists():
                return path

        candidate = self.settings.backend_root / "data" / "catalog_bundle" / "products.json"
        if candidate.exists():
            return candidate
        return None

    def _discover_workbook_path(self) -> Path:
        if self.settings.product_xlsx_path:
            return Path(self.settings.product_xlsx_path)

        candidates = sorted(
            path
            for path in self.settings.project_root.glob("*.xlsx")
            if not path.name.startswith("~$")
        )
        if not candidates:
            raise FileNotFoundError("未找到 Excel 货盘文件。")
        return candidates[0]

    def load_catalog(self, force_reload: bool = False) -> None:
        if self.loaded and not force_reload:
            return

        json_path = self._discover_json_path()
        if json_path is not None:
            self._load_from_json(json_path)
            return

        workbook_path = self._discover_workbook_path()
        self.asset_url_by_image_id = self._extract_cell_images(workbook_path)

        wb = load_workbook(workbook_path, read_only=True, data_only=True)
        ws = wb[wb.sheetnames[0]]

        rows = ws.iter_rows(values_only=True)
        headers = [_normalize_header(item) for item in next(rows)]

        products: list[dict[str, Any]] = []
        for row in rows:
            item = {headers[idx]: row[idx] for idx in range(min(len(headers), len(row)))}
            code = _clean_text(item.get("商品编码"))
            name = _clean_text(item.get("商品名称"))
            if not code or not name:
                continue

            products.append(
                {
                    "product_name": name,
                    "product_code": code,
                    "product_image_url": self._resolve_media_field(item.get("商品图片")),
                    "product_qr_url": self._resolve_media_field(item.get("商品二维码")),
                    "selling_points": _clean_text(item.get("产品优势卖点")),
                    "main_material": _clean_text(item.get("主材质")),
                    "stone_material": _clean_text(item.get("配石材质")),
                    "group_price": _safe_float(item.get("开团价")),
                    "wholesale_price": self._resolve_wholesale_price(item),
                    "discount": _safe_float(item.get("折扣")),
                    "system_category": _clean_text(item.get("系统款式")),
                    "system_attributes": _clean_text(item.get("系统属性")),
                    "suitable_people": _clean_text(item.get("适合人群")),
                    "luxury_flag": _clean_text(item.get("显贵款")),
                }
            )

        self.workbook_path = workbook_path
        self.catalog_source_path = workbook_path
        self.products = self._suppress_duplicate_qr_assets(products)
        self.loaded = True

    def summary(self) -> dict[str, Any]:
        categories = sorted(
            {
                product["system_category"]
                for product in self.products
                if product.get("system_category")
            }
        )
        return {
            "loaded": self.loaded,
            "workbook_path": str(self.catalog_source_path or self.workbook_path) if (self.catalog_source_path or self.workbook_path) else None,
            "product_count": len(self.products),
            "categories": categories,
        }

    def _load_from_json(self, json_path: Path) -> None:
        payload = json.loads(json_path.read_text(encoding="utf-8"))
        if not isinstance(payload, list):
            raise ValueError("products.json 格式不正确，顶层必须是数组。")

        products: list[dict[str, Any]] = []
        for item in payload:
            if not isinstance(item, dict):
                continue
            code = _clean_text(item.get("product_code"))
            name = _clean_text(item.get("product_name"))
            if not code or not name:
                continue

            products.append(
                {
                    "product_name": name,
                    "product_code": code,
                    "product_image_url": self._normalize_static_path(item.get("product_image_url")),
                    "product_qr_url": self._normalize_static_path(item.get("product_qr_url")),
                    "selling_points": _clean_text(item.get("selling_points")),
                    "main_material": _clean_text(item.get("main_material")),
                    "stone_material": _clean_text(item.get("stone_material")),
                    "group_price": _safe_float(item.get("group_price")),
                    "wholesale_price": _safe_float(item.get("wholesale_price")),
                    "discount": _safe_float(item.get("discount")),
                    "system_category": _clean_text(item.get("system_category")),
                    "system_attributes": _clean_text(item.get("system_attributes")),
                    "suitable_people": _clean_text(item.get("suitable_people")),
                    "luxury_flag": _clean_text(item.get("luxury_flag")),
                }
            )

        self.workbook_path = None
        self.catalog_source_path = json_path
        self.products = self._suppress_duplicate_qr_assets(products)
        self.loaded = True

    def _suppress_duplicate_qr_assets(self, products: list[dict[str, Any]]) -> list[dict[str, Any]]:
        seen_hashes: dict[str, str] = {}
        normalized: list[dict[str, Any]] = []
        for product in products:
            item = dict(product)
            qr_path = self._resolve_static_media_path(item.get("product_qr_url"))
            if qr_path is not None:
                digest = hashlib.sha256(qr_path.read_bytes()).hexdigest()
                if digest in seen_hashes:
                    item["product_qr_url"] = None
                else:
                    seen_hashes[digest] = str(item.get("product_code") or "")
            normalized.append(item)
        return normalized

    def _resolve_static_media_path(self, value: Any) -> Path | None:
        text = str(value or "").strip()
        if not text.startswith("/static/"):
            return None
        relative = text.removeprefix("/static/").replace("/", "\\")
        local_path = self.settings.backend_root / "data" / Path(relative)
        if not local_path.exists():
            return None
        return local_path

    def _normalize_static_path(self, value: Any) -> str | None:
        text = str(value or "").strip()
        if not text:
            return None
        if text.startswith(("http://", "https://", "/static/")):
            return text
        return None

    def _resolve_wholesale_price(self, item: dict[str, Any]) -> float | None:
        wholesale_price = _safe_float(item.get("批发裸价"))
        if wholesale_price is not None:
            return wholesale_price

        group_price = _safe_float(item.get("开团价"))
        discount = _safe_float(item.get("折扣"))
        if group_price is not None and discount is not None:
            return round(group_price * discount, 2)
        return None

    def _resolve_media_field(self, value: Any) -> str | None:
        text = str(value or "").strip()
        if not text:
            return None
        if text.startswith("http://") or text.startswith("https://") or text.startswith("/static/"):
            return text

        image_id = _extract_dispimg_id(text)
        if image_id:
            return self.asset_url_by_image_id.get(image_id)
        return None

    def _extract_cell_images(self, workbook_path: Path) -> dict[str, str]:
        asset_dir = self.settings.backend_root / "data" / "catalog_media"
        asset_dir.mkdir(parents=True, exist_ok=True)

        with ZipFile(workbook_path) as archive:
            try:
                rels_root = ET.fromstring(archive.read("xl/_rels/cellimages.xml.rels"))
                cellimages_root = ET.fromstring(archive.read("xl/cellimages.xml"))
            except KeyError:
                return {}

            rel_ns = {"r": "http://schemas.openxmlformats.org/package/2006/relationships"}
            cell_ns = {
                "etc": "http://www.wps.cn/officeDocument/2017/etCustomData",
                "xdr": "http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing",
                "a": "http://schemas.openxmlformats.org/drawingml/2006/main",
            }

            rel_target_by_id: dict[str, str] = {}
            for rel in rels_root.findall("r:Relationship", rel_ns):
                rel_id = rel.attrib.get("Id")
                target = rel.attrib.get("Target")
                if rel_id and target:
                    rel_target_by_id[rel_id] = target

            asset_url_by_image_id: dict[str, str] = {}
            for cell_image in cellimages_root.findall("etc:cellImage", cell_ns):
                c_nv_pr = cell_image.find(".//xdr:cNvPr", cell_ns)
                blip = cell_image.find(".//a:blip", cell_ns)
                if c_nv_pr is None or blip is None:
                    continue

                image_id = c_nv_pr.attrib.get("name")
                rel_id = blip.attrib.get("{http://schemas.openxmlformats.org/officeDocument/2006/relationships}embed")
                target = rel_target_by_id.get(rel_id or "")
                if not image_id or not target:
                    continue

                archive_path = f"xl/{target.lstrip('./')}"
                suffix = Path(target).suffix or ".png"
                output_name = f"{image_id}{suffix}"
                output_path = asset_dir / output_name

                if not output_path.exists():
                    try:
                        output_path.write_bytes(archive.read(archive_path))
                    except KeyError:
                        continue

                asset_url_by_image_id[image_id] = f"/static/catalog_media/{output_name}"

            return asset_url_by_image_id
